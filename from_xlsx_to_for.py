import openpyxl

# Step 1: Open the Excel file 
excel_file = 'Lap_data.xlsx'
wb = openpyxl.load_workbook(excel_file)

wb_s = wb['Speed_values']
wb_b = wb['Brake_intensity']

insert_lines_b = []
insert_lines_s = []
flag_writting = 0

# Step 2: Modify the Fortran file
fortran_file = input("Which Fortran file ? (With .for at the end of the name and without brackets)")

######### FOR THE BRAKE #########

with open(fortran_file, 'r') as file:
    lines = file.readlines()

    # Find the first line starting with 'r' and insert fortran lines after it
    for i, line in enumerate(lines):
        if (line.strip().lower().startswith('!b')  and (flag_writting == 0) ):
            if (lines[i+1].strip().lower().startswith('!r')):
                print("File not found or there is already a brakeblock")
                Answer = input("Do we earase the brakeblock ?(Y/n)")
                if Answer == "Y":
                    k=i+1
                    while ((lines[k].strip().lower().startswith('!/')) == False):
                        del lines[k]
                    print("Brake block will be replaced with the new one.")
                else:
                    print("Previous brakeblock intact")
                
            if not(lines[i+1].strip().lower().startswith('!r')):
                insert_lines_b.append('      !r\n')
                for j in range(7,wb_b.max_row):
                    if wb_b['A' + str(j)].value == 'neutral' :
                        insert_lines_b.append('      IF((TIME(2) .LE. '+ str(wb_b['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_b['B' + str(j)].value) + ')) THEN')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('        brake_i = 0')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('      ENDIF')       
                        insert_lines_b.append('\n')                        
                    if wb_b['A' + str(j)].value == 'start' :
                        insert_lines_b.append('      IF((TIME(2) .LE. '+ str(wb_b['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_b['B' + str(j)].value) + ')) THEN')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('        brake_i = 1')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('      ENDIF')
                        insert_lines_b.append('\n')
                    if wb_b['A' + str(j)].value == 'middle' :
                        insert_lines_b.append('      IF((TIME(2) .LE. '+ str(wb_b['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_b['B' + str(j)].value) + ')) THEN')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('        brake_i = '+ str(wb_b['D' + str(j+1)].value) + '*TIME(2) + ' + str(-wb_b['D' + str(j+1)].value*wb_b['B' + str(j+1)].value))
                        insert_lines_b.append('\n')
                        insert_lines_b.append('      ENDIF')
                        insert_lines_b.append('\n')
                    if wb_b['A' + str(j)].value == 'end' :
                        insert_lines_b.append('      IF((TIME(2) .LE. '+ str(wb_b['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_b['B' + str(j)].value) + ')) THEN')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('        brake_i = 0')
                        insert_lines_b.append('\n')
                        insert_lines_b.append('      ENDIF')
                        insert_lines_b.append('\n')
                    flag_writting = 1
                print("Brake Block added")
                lines[i + 1:i + 1] = insert_lines_b

    with open(fortran_file, 'w') as file:
        file.writelines(lines)
        
######### FOR THE SPEED AND THE ANGLE #########

with open(fortran_file, 'r') as file:
    lines = file.readlines()

    # Find the first line starting with 'r' and insert fortran lines after it
    for i, line in enumerate(lines):
        if (line.strip().lower().startswith('!s')):
            if (lines[i+1].strip().lower().startswith('!r')):
                print("File not found or there is already a speedblock")
                Answer = input("Do we earase the speedblock ?(Y/n)")
                if Answer == "Y":
                    k=i+1
                    while ((lines[k].strip().lower().startswith('!/')) == False):
                        del lines[k]
                    print("Speed block will be replaced with the new one.")
                else:
                    print("Previous speedblock intact")
                
            if not (lines[i+1].strip().lower().startswith('!r')):
                insert_lines_s.append('      !r\n')
                insert_lines_s.append('      do j=1,2\n')
                for j in range(7,wb_s.max_row):
                    if wb_s['A' + str(j)].value == 'acceleration' :
                        insert_lines_s.append('        IF((TIME(2) .LE. '+ str(wb_s['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_s['B' + str(j)].value) + ')) THEN')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             speed_i = ' + str(wb_s['D' + str(j)].value) + '*TIME(2) + (' + str(wb_s['C' + str(j)].value-wb_s['D' + str(j)].value*wb_s['B' + str(j)].value)+')')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((speed_i - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6)')  
                        insert_lines_s.append('\n')                        
                        insert_lines_s.append('             exit')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ELSE')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((' + str(wb_s['C' + str(j+1)].value) + ' - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6/1000)')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ENDIF')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('\n')
                    if wb_s['A' + str(j)].value == 'deceleration' :
                        insert_lines_s.append('        IF((TIME(2) .LE. '+ str(wb_s['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_s['B' + str(j)].value) + ')) THEN')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             speed_i = ' + str(wb_s['D' + str(j)].value) + '*TIME(2) + (' + str(wb_s['C' + str(j)].value-wb_s['D' + str(j)].value*wb_s['B' + str(j)].value)+')')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((speed_i - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6/1000)')  
                        insert_lines_s.append('\n')                        
                        insert_lines_s.append('             exit')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ELSE')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((' + str(wb_s['C' + str(j+1)].value) + ' - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6/1000)') 
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ENDIF')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('\n')
                    if wb_s['A' + str(j)].value == 'neutral' :
                        insert_lines_s.append('        IF((TIME(2) .LE. '+ str(wb_s['B' + str(j+1)].value)+') .AND. (TIME(2) .GT. '+ str(wb_s['B' + str(j)].value) + ')) THEN')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             speed_i = ' + str(wb_s['C' + str(j)].value))
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((speed_i - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6/1000)')  
                        insert_lines_s.append('\n')                        
                        insert_lines_s.append('             exit')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ELSE')  
                        insert_lines_s.append('\n')
                        insert_lines_s.append('             angle_i = angle_i + abs(((' + str(wb_s['C' + str(j+1)].value) + ' - ' + str(wb_s['C' + str(j)].value) + ')/2) + 35)/(tire_radius*3.6/1000)') 
                        insert_lines_s.append('\n')
                        insert_lines_s.append('        ENDIF')
                        insert_lines_s.append('\n')
                        insert_lines_s.append('\n')
                insert_lines_s.append('      end do\n')
                print("Speed Block added")
                lines[i + 1:i + 1] = insert_lines_s

    with open(fortran_file, 'w') as file:
        file.writelines(lines)
